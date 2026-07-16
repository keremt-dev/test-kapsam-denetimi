#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
test-kapsam-denetimi · Adım 4: Sentez (yalnız verify PASS sonrası)
result JSON'lar + manifest -> 4 çıktı. TÜM sayımlar madde durumundan hesaplanır.
  <MOD>_Kapsam_Bosluk_Raporu.md
  <MOD>_Kapsam_Bosluk_Matrisi.xlsx  (Matris / Ozet / Etiketsiz_Artik)
  <test dosyası> - Inceleme.xlsx     (orijinalin kopyası + 'Kapsam İnceleme Notu' sütunu)
  <MOD>_jira_yorumlari.md            (kapsam-içi oran; Yok + önemli Kısmi)

Kullanım: python synthesize.py --workdir work --outdir .
"""
import argparse, glob, json, os, re, shutil, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TURORD = {'Ön/Son Koşul': 0, 'Ana Akış': 1, 'Alternatif': 2, 'İş Kuralı': 3, 'SRS Gereksinimi': 4}

def natkey(s):
    """Doğal sıralama anahtarı: A2 < A10."""
    return [int(p) if p.isdigit() else p for p in re.split(r'(\d+)', str(s))]

def madde_sort(maddeler):
    return sorted(maddeler, key=lambda x: (TURORD.get(x['tur'], 9), natkey(x['no'])))

def load(workdir):
    man = json.load(open(os.path.join(workdir, 'manifest.json'), encoding='utf-8'))
    data, order = {}, []
    for f in sorted(glob.glob(os.path.join(workdir, '*_result.json'))):
        d = json.load(open(f, encoding='utf-8'))
        data[d['ks']] = d; order.append(d['ks'])
    order.sort()
    return man, data, order

def counts(d):
    c = collections.Counter(m['durum'] for m in d['maddeler'])
    dropped = sum(1 for m in d['maddeler'] if m['durum'] == 'Kısmi' and m['tur'] == 'Ön/Son Koşul')
    return c['Tam'], c['Kısmi'], c['Yok'], len(d['maddeler']), dropped

def sj(s):
    return ' '.join((s or '').split())

def md(s):
    """Markdown tablo hücresi: boşluk normalize + '|' kaçır."""
    return sj(s).replace('|', '\\|')

# ---------------- 1) MD RAPOR ----------------
def build_report(man, data, order, outdir):
    mod = man['module']
    n_tests = sum(len(uc.get('valid_test_ids', [])) for uc in man['use_cases'].values())
    L = [f"# {mod} Modülü — Test ↔ Kullanım Senaryosu Kapsam Boşluk Raporu\n",
         f"**Kapsam:** {len(order)} kullanım senaryosu · {n_tests} test senaryosu",
         "**Yöntem:** Her KS'nin ana akışı, alternatifleri, iş kuralları, ön/son koşulları ve ilgili SRS gereksinimleri "
         "atomik maddelere ayrılıp Tam/Kısmi/Yok olarak eşleştirildi. Her KS ayrı Opus 4.8 ajanı ile denetlendi.\n", "---\n"]
    tt = collections.Counter()
    for k in order:
        tam, kis, yok, tot, _ = counts(data[k]); tt['tam'] += tam; tt['kis'] += kis; tt['yok'] += yok; tt['tot'] += tot
    L.append("## 1. Yönetici Özeti\n")
    L.append(f"Toplam **{tt['tot']} madde**: **{tt['tam']} Tam** (%{100*tt['tam']/tt['tot']:.0f}), "
             f"**{tt['kis']} Kısmi** (%{100*tt['kis']/tt['tot']:.0f}), **{tt['yok']} Yok** (%{100*tt['yok']/tt['tot']:.0f}).\n")
    L.append("| KS | Ad | Madde | Tam | Kısmi | Yok | Tam-Kapsam % |")
    L.append("|----|----|------:|----:|------:|----:|----:|")
    for k in order:
        tam, kis, yok, tot, _ = counts(data[k])
        L.append(f"| {k} | {data[k].get('ks_adi','')} | {tot} | {tam} | {kis} | {yok} | %{100*tam/tot:.0f} |")
    L.append(f"| **TOPLAM** |  | **{tt['tot']}** | **{tt['tam']}** | **{tt['kis']}** | **{tt['yok']}** | **%{100*tt['tam']/tt['tot']:.0f}** |\n")

    L.append("## 2. Kritik Boşluklar — Test Edilmeyen Maddeler (Yok)\n")
    L.append("| KS | Tür | Madde | Özet | Boşluk Notu |"); L.append("|----|----|----|----|----|")
    for k in order:
        for m in data[k]['maddeler']:
            if m['durum'] == 'Yok':
                L.append(f"| {k} | {m['tur']} | {m['no']} | {md(m['ozet'])} | {md(m.get('not',''))} |")
    L.append("")
    L.append("## 3. Eksik Kapsam — Kısmen Test Edilen Maddeler\n")
    L.append("| KS | Tür | Madde | Özet | Karşılayan Test | Eksik Yön |"); L.append("|----|----|----|----|----|----|")
    for k in order:
        for m in data[k]['maddeler']:
            if m['durum'] == 'Kısmi':
                L.append(f"| {k} | {m['tur']} | {m['no']} | {md(m['ozet'])} | {', '.join(m.get('karsilayan_testler',[]))} | {md(m.get('not',''))} |")
    L.append("")
    L.append("## 4. Ters Yön Bulguları\n### 4.1. Gereksinim Etiketi Boş Testler\n")
    L.append("| KS | Test ID | Test Adı |"); L.append("|----|----|----|")
    etot = 0
    for k in order:
        for t in data[k].get('etiketsiz_testler', []):
            L.append(f"| {k} | {t['test_id']} | {t.get('ad','')} |"); etot += 1
    L.append(f"\n_Toplam {etot} etiketsiz test._\n")
    L.append("### 4.2. Artık / İskelet Test Satırları\n")
    arows = [(k, t) for k in order for t in data[k].get('artik_testler', [])]
    if arows:
        L.append("| KS | Test ID | Başlık | Not |"); L.append("|----|----|----|----|")
        for k, t in arows:
            L.append(f"| {k} | {t['test_id']} | {md(t.get('ad',''))} | {md(t.get('not',''))} |")
    else:
        L.append("_Artık test bulunmadı._")
    L.append(f"\n_Toplam {len(arows)} artık/iskelet satır._\n\n---\n\n## 5. Kullanım Senaryosu Bazında Detay\n")
    for k in order:
        tam, kis, yok, tot, _ = counts(data[k])
        L.append(f"### {k} — {data[k].get('ks_adi','')}\n")
        L.append(f"Madde: {tot} · Tam: {tam} · Kısmi: {kis} · Yok: {yok}\n")
        L.append("| Tür | No | Özet | Karşılayan Test | Durum | Not |"); L.append("|----|----|----|----|----|----|")
        for m in madde_sort(data[k]['maddeler']):
            L.append(f"| {m['tur']} | {m['no']} | {md(m['ozet'])} | {', '.join(m.get('karsilayan_testler',[])) or '—'} | {m['durum']} | {md(m.get('not',''))} |")
        L.append("")
    p = os.path.join(outdir, f"{mod}_Kapsam_Bosluk_Raporu.md")
    open(p, 'w', encoding='utf-8').write('\n'.join(L))
    return p

# ---------------- 2) MATRİS XLSX ----------------
def build_matrix(man, data, order, outdir):
    mod = man['module']
    wb = openpyxl.Workbook()
    hf = PatternFill('solid', fgColor='1F4E78'); hfont = Font(bold=True, color='FFFFFF')
    red = PatternFill('solid', fgColor='F8CBAD'); yel = PatternFill('solid', fgColor='FFE699'); grn = PatternFill('solid', fgColor='C6EFCE')
    thin = Border(*[Side(style='thin', color='BFBFBF')] * 4); wrap = Alignment(wrap_text=True, vertical='top')
    df = {'Yok': red, 'Kısmi': yel, 'Tam': grn}
    ws = wb.active; ws.title = 'Matris'
    cols = ['KS', 'Tür', 'Madde No', 'Madde Özeti', 'Karşılayan Test ID', 'Durum', 'Boşluk / Not']
    ws.append(cols)
    for c in range(1, len(cols)+1):
        ws.cell(1, c).fill = hf; ws.cell(1, c).font = hfont
    r = 2
    for k in order:
        for m in madde_sort(data[k]['maddeler']):
            ws.append([k, m['tur'], m['no'], sj(m['ozet']), ', '.join(m.get('karsilayan_testler', [])) or '—', m['durum'], sj(m.get('not', ''))])
            ws.cell(r, 6).fill = df.get(m['durum']); ws.cell(r, 6).font = Font(bold=True)
            for c in range(1, 8):
                ws.cell(r, c).border = thin; ws.cell(r, c).alignment = wrap
            r += 1
    for i, w in enumerate([9, 15, 10, 46, 26, 9, 50], 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = 'A2'; ws.auto_filter.ref = f"A1:G{r-1}"

    ws2 = wb.create_sheet('Ozet')
    ws2.append(['KS', 'Ad', 'Madde', 'Tam', 'Kısmi', 'Yok', 'Tam-Kapsam %'])
    for c in range(1, 8): ws2.cell(1, c).fill = hf; ws2.cell(1, c).font = hfont
    tt = collections.Counter(); rr = 2
    for k in order:
        tam, kis, yok, tot, _ = counts(data[k]); tt['tam'] += tam; tt['kis'] += kis; tt['yok'] += yok; tt['tot'] += tot
        ws2.append([k, data[k].get('ks_adi', ''), tot, tam, kis, yok, round(100*tam/tot)]); rr += 1
    ws2.append(['TOPLAM', '', tt['tot'], tt['tam'], tt['kis'], tt['yok'], round(100*tt['tam']/tt['tot'])])
    for c in range(1, 8): ws2.cell(rr, c).font = Font(bold=True)
    for i, w in enumerate([9, 40, 8, 7, 8, 7, 14], 1): ws2.column_dimensions[chr(64+i)].width = w

    ws3 = wb.create_sheet('Etiketsiz_Artik')
    ws3.append(['Tür', 'KS', 'Test ID', 'Test Adı / Not'])
    for c in range(1, 5): ws3.cell(1, c).fill = hf; ws3.cell(1, c).font = hfont
    for k in order:
        for t in data[k].get('etiketsiz_testler', []):
            ws3.append(['Etiketsiz (gereksinim boş)', k, t['test_id'], t.get('ad', '')])
    for k in order:
        for t in data[k].get('artik_testler', []):
            ws3.append(['İskelet/Artık satır', k, t['test_id'], t.get('ad', '')])
    for i, w in enumerate([26, 9, 24, 60], 1): ws3.column_dimensions[chr(64+i)].width = w
    ws3.auto_filter.ref = f"A1:D{ws3.max_row}"

    p = os.path.join(outdir, f"{mod}_Kapsam_Bosluk_Matrisi.xlsx")
    try:
        wb.save(p)
        return p, None
    except PermissionError:
        return p, f"UYARI: {os.path.basename(p)} açık/kilitli, kaydedilemedi. Excel'de kapatıp tekrar çalıştırın."

# ---------------- 3) YORUMLU TEST KOPYASI ----------------
def build_annotated(man, data, order, outdir):
    src = man['tests_file']
    base = os.path.splitext(os.path.basename(src))[0]
    dst = os.path.join(outdir, f"{base} - Inceleme.xlsx")
    notemap = {}
    for k in order:
        for tn in data[k].get('test_notlari', []):
            notemap[tn['test_id']] = tn
    hf = PatternFill('solid', fgColor='1F4E78'); hfont = Font(bold=True, color='FFFFFF')
    wrap = Alignment(wrap_text=True, vertical='top')
    try:
        shutil.copy(src, dst)
    except PermissionError:
        return dst, f"UYARI: {os.path.basename(dst)} açık/kilitli, kopyalanamadı."
    wb = openpyxl.load_workbook(dst)
    filled = 0
    for sh in wb.worksheets:
        last = 1
        for c in range(1, sh.max_column + 1):
            if sh.cell(1, c).value not in (None, ''):
                last = c
        nc = last + 1
        hc = sh.cell(1, nc); hc.value = 'Kapsam İnceleme Notu'; hc.fill = hf; hc.font = hfont
        for row in range(2, sh.max_row + 1):
            tid = sh.cell(row, 1).value
            if not tid:
                continue
            tid = str(tid).strip()
            if tid.lower() == 'id':  # tekrar eden başlık satırı — nota gerek yok
                continue
            n = notemap.get(tid)
            cell = sh.cell(row, nc); cell.alignment = wrap
            if n:
                cell.value = f"Karşıladığı: {n.get('karsiladigi','')} | Yeterlilik: {n.get('yeterlilik','')} | {n.get('yorum','')}"
                cell.font = Font(color='1F4E78'); filled += 1
            else:
                cell.value = '(iskelet/taslak satır — adım ve beklenen sonuç boş; tamamlanmalı veya silinmeli)'
                cell.font = Font(color='C00000', italic=True)
    try:
        wb.save(dst)
        return dst, None
    except PermissionError:
        return dst, f"UYARI: {os.path.basename(dst)} kaydedilemedi (kilitli)."

# ---------------- 4) JIRA YORUMLARI ----------------
def build_jira(man, data, order, outdir):
    mod = man['module']
    L = [f"# {mod} — KS Bazında Test Kapsam Notları", "",
         "İlgili kullanım senaryosunun Jira issue'suna yorum olarak eklenmek üzere hazırlandı. "
         "Düşük öncelikli ön koşul varsayımları (login, entegrasyonun hazır olması vb.) kapsam dışı bırakıldı; "
         "oranlar yalnızca ele alınan maddeleri sayar.", ""]
    for k in order:
        d = data[k]; tam, kis, yok, tot, dropped = counts(d); scope = tot - dropped
        yoks = [m for m in d['maddeler'] if m['durum'] == 'Yok']
        impk = [m for m in d['maddeler'] if m['durum'] == 'Kısmi' and m['tur'] != 'Ön/Son Koşul']
        etk = [t['test_id'] for t in d.get('etiketsiz_testler', [])]
        artik = d.get('artik_testler', [])
        L.append(f"## {k} — {d.get('ks_adi','')}"); L.append("")
        if not yoks and not impk:
            L.append(f"Ele alınan {scope} maddenin tamamına yakını test ediliyor; önemli bir boşluk yok.")
        else:
            L.append(f"Ele alınan {scope} maddenin {tam} tanesi tam karşılanıyor. Eksikler:"); L.append("")
            for m in yoks:
                L.append(f"- **{m['no']}** — {sj(m.get('not',''))} Karşılayan test yok, eklenmeli.")
            for m in impk:
                t = ', '.join(m.get('karsilayan_testler', []))
                ref = f" ({t} var ama yetersiz)" if t else ""
                L.append(f"- **{m['no']}** — {sj(m.get('not',''))}{ref}")
        tail = []
        if artik:
            ids = ', '.join(t['test_id'] for t in artik[:3]) + (' …' if len(artik) > 3 else '')
            tail.append(f"Sayfada yalnızca başlık içeren {len(artik)} yarım test satırı ({ids}) var; tamamlanmalı ya da silinmeli.")
        if etk:
            tail.append("Gereksinim sütunu boş testler (SRS bağı eklenmeli): " + ", ".join(etk) + ".")
        if tail:
            L.append(""); L.append(" ".join(tail))
        L.append("")
    p = os.path.join(outdir, f"{mod}_jira_yorumlari.md")
    open(p, 'w', encoding='utf-8').write('\n'.join(L))
    return p

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--workdir', default='work')
    ap.add_argument('--outdir', default='.')
    a = ap.parse_args()
    man, data, order = load(a.workdir)
    os.makedirs(a.outdir, exist_ok=True)
    warns = []
    p1 = build_report(man, data, order, a.outdir); print("rapor :", p1)
    p2, w2 = build_matrix(man, data, order, a.outdir); print("matris:", p2); warns += [w2] if w2 else []
    p3, w3 = build_annotated(man, data, order, a.outdir); print("kopya :", p3); warns += [w3] if w3 else []
    p4 = build_jira(man, data, order, a.outdir); print("jira  :", p4)
    for w in warns:
        print("  " + w)

if __name__ == '__main__':
    main()
