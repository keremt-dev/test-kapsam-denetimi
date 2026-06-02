#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
test-kapsam-denetimi · Faz 3: SERT DOĞRULAMA (sentezden önce kapı)
result JSON'larını + manifest'i okur, tutarlılık kontrolleri yapar.
- C1 sayım bütünlüğü: ozet'i madde durumundan YENİDEN HESAPLAR ve dosyaya yazar
- C2 uydurma yok: atıf yapılan her Test ID manifest'te var
- C3 ters kapsam tam: her geçerli Test ID test_notlari'nda tam bir kez
- C4 SRS/KS: KS'nin SRS-boyutu == manifest KS SRS listesi
- C5 modül SRS bütünlüğü: tüm KS SRS-boyutu birleşimi == modül SRS evreni
- C6 enum: durum / tur geçerli

Herhangi bir FAIL -> exit code 1 (sentez yapılmamalı).
--post modu: sentez sonrası çıktı dosyalarını kontrol eder.

Kullanım:
  python verify.py --workdir work
  python verify.py --workdir work --post --outdir <ciktilarin_oldugu_dizin>
"""
import argparse, glob, json, os, re, sys, collections

DURUMS = {'Tam', 'Kısmi', 'Yok'}
TURS = {'Ana Akış', 'Alternatif', 'İş Kuralı', 'SRS Gereksinimi', 'Ön/Son Koşul'}

def load(workdir):
    man = json.load(open(os.path.join(workdir, 'manifest.json'), encoding='utf-8'))
    results = {}
    for f in sorted(glob.glob(os.path.join(workdir, '*_result.json'))):
        d = json.load(open(f, encoding='utf-8'))
        results[d['ks']] = (f, d)
    return man, results

def srs_tail(s):
    m = re.findall(r'\d+', s)
    return m[-1] if m else s

def verify(workdir):
    man, results = load(workdir)
    fails = []
    ucs = man['use_cases']

    # her KS sonuçları var mı
    for ks in ucs:
        if ks not in results:
            fails.append(f"[EKSİK] {ks} için result.json yok")

    srs_seen = set()
    for ks, (path, d) in results.items():
        info = ucs.get(ks, {})
        valid = set(info.get('valid_test_ids', []))
        maddeler = d.get('maddeler', [])

        # C6 enum
        for m in maddeler:
            if m.get('durum') not in DURUMS:
                fails.append(f"[C6 {ks}] geçersiz durum: {m.get('no')}={m.get('durum')}")
            if m.get('tur') not in TURS:
                fails.append(f"[C6 {ks}] geçersiz tur: {m.get('no')}={m.get('tur')}")

        # C1 sayım bütünlüğü + ozet yeniden yaz
        cnt = collections.Counter(m.get('durum') for m in maddeler)
        recomputed = {'tam': cnt['Tam'], 'kısmi': cnt['Kısmi'], 'yok': cnt['Yok'],
                      'toplam_madde': len(maddeler)}
        if d.get('ozet') != recomputed:
            d['ozet'] = recomputed
            json.dump(d, open(path, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
            print(f"[C1 {ks}] ozet yeniden hesaplandı -> {recomputed}")

        # C2 uydurma yok
        for m in maddeler:
            for tid in m.get('karsilayan_testler', []):
                if tid not in valid:
                    fails.append(f"[C2 {ks}] uydurma/geçersiz Test ID (madde {m.get('no')}): {tid}")
        for tn in d.get('test_notlari', []):
            if tn.get('test_id') not in valid:
                fails.append(f"[C2 {ks}] test_notlari geçersiz Test ID: {tn.get('test_id')}")

        # C3 ters kapsam tam
        noted = collections.Counter(tn.get('test_id') for tn in d.get('test_notlari', []))
        for tid in valid:
            if noted[tid] == 0:
                fails.append(f"[C3 {ks}] test_notu eksik: {tid}")
            elif noted[tid] > 1:
                fails.append(f"[C3 {ks}] test_notu mükerrer: {tid} ({noted[tid]}x)")

        # C4 SRS/KS
        ks_srs_dim = {srs_tail(m['no']) for m in maddeler if m.get('tur') == 'SRS Gereksinimi'}
        ks_srs_exp = {srs_tail(s) for s in info.get('srs', [])}
        if ks_srs_dim != ks_srs_exp:
            miss = ks_srs_exp - ks_srs_dim; extra = ks_srs_dim - ks_srs_exp
            fails.append(f"[C4 {ks}] SRS-boyutu uyuşmuyor. Eksik={sorted(miss)} Fazla={sorted(extra)}")
        srs_seen |= ks_srs_dim

    # C5 modül SRS bütünlüğü
    universe = {srs_tail(s) for s in man.get('srs_universe', [])}
    if srs_seen != universe:
        miss = universe - srs_seen; extra = srs_seen - universe
        if miss or extra:
            fails.append(f"[C5] Modül SRS kapsamı eksik. Testlerde yok={sorted(miss)} Fazla={sorted(extra)}")

    return fails

def verify_post(workdir, outdir):
    import openpyxl
    man, results = load(workdir)
    fails = []
    total_madde = sum(len(d.get('maddeler', [])) for _, d in results.values())
    mod = man['module']
    mx = os.path.join(outdir, f"{mod}_Kapsam_Bosluk_Matrisi.xlsx")
    if not os.path.exists(mx):
        fails.append(f"[POST] matris bulunamadı: {mx}")
        return fails
    wb = openpyxl.load_workbook(mx)
    ms = wb['Matris']
    rows = ms.max_row - 1
    if rows != total_madde:
        fails.append(f"[POST] Matris satır {rows} != Σ madde {total_madde}")
    # Ozet TOPLAM kontrolü
    oz = wb['Ozet']
    tam = sum(collections.Counter(m['durum'] for m in d['maddeler'])['Tam'] for _, d in results.values())
    last = [oz.cell(oz.max_row, c).value for c in range(1, 8)]
    if last[3] != tam:
        fails.append(f"[POST] Ozet TOPLAM tam {last[3]} != {tam}")
    return fails

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--workdir', default='work')
    ap.add_argument('--post', action='store_true')
    ap.add_argument('--outdir', default='.')
    a = ap.parse_args()
    fails = verify_post(a.workdir, a.outdir) if a.post else verify(a.workdir)
    print('\n' + '=' * 60)
    if fails:
        print(f"DOĞRULAMA: FAIL ({len(fails)} sorun)")
        for f in fails:
            print('  ✗ ' + f)
        sys.exit(1)
    print("DOĞRULAMA: PASS — tüm kontroller geçti.")
    sys.exit(0)

if __name__ == '__main__':
    main()
