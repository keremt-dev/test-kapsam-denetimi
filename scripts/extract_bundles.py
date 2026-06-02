#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
test-kapsam-denetimi · Faz 0: Çıkarım
3 kaynak dokümandan (SRS .xlsx, kullanım senaryoları .docx, test senaryoları .xlsx)
her kullanım senaryosu için kendine-yeten bir girdi paketi (<KS>_input.md) ve
bir manifest.json üretir.

Tüm kolon/tablo tespiti BAŞLIK-TABANLIDIR (sıra indeksine güvenmez).

Kullanım:
  python extract_bundles.py --indir <klasor> [--outdir work] [--module KEY]
  python extract_bundles.py --reqs srs.xlsx --usecases uc.docx --tests test.xlsx --outdir work
"""
import argparse, json, os, re, sys, glob
import openpyxl
import docx

# ---------- yardımcılar ----------
def norm(s):
    s = '' if s is None else str(s)
    return re.sub(r'\s+', ' ', s.strip()).lower().replace('ı', 'i').replace('İ', 'i')

def find_header_row(ws, must_have):
    """İlk 5 satırda, verilen başlıklardan birini içeren satırı başlık satırı kabul et."""
    want = {norm(x) for x in must_have}
    for r in range(1, min(ws.max_row, 6) + 1):
        vals = {norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if want & vals:
            return r
    return 1

def col_map(ws, header_row):
    """Başlık adı(normalize) -> kolon indeksi."""
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip():
            m[norm(v)] = c
    return m

def pick(cmap, *aliases):
    for a in aliases:
        if norm(a) in cmap:
            return cmap[norm(a)]
    return None

def ks_short(ksno):
    """BVA_KP_KEY_KS_001 -> KS_001"""
    mt = re.search(r'(KS[_\-]?\d+)', str(ksno), re.I)
    return mt.group(1).upper().replace('-', '_') if mt else str(ksno).strip()

def module_of(ksno):
    """BVA_KP_KEY_KS_001 -> KEY"""
    mt = re.search(r'BVA[_\-]?KP[_\-]?(.+?)[_\-]?KS[_\-]?\d+', str(ksno), re.I)
    return mt.group(1).strip('_-').upper() if mt else None

# ---------- dosya sınıflandırma ----------
OUTPUT_MARKERS = ('inceleme', 'kapsam_bosluk', 'kapsam bosluk', 'matris', '_jira_')
def is_output(f):
    b = os.path.basename(f).lower()
    return any(m in b for m in OUTPUT_MARKERS)

def classify(indir):
    docx_f = reqs_f = None
    tests_cands = []
    for f in glob.glob(os.path.join(indir, '*')):
        low = f.lower()
        if os.path.basename(f).startswith('~') or is_output(f):
            continue
        if low.endswith('.docx'):
            docx_f = docx_f or f
        elif low.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            except Exception:
                continue
            # SRS: 'Gereksinim Numarası' başlığı var mı
            is_srs = False
            ws0 = wb[wb.sheetnames[0]]
            for r in range(1, min(ws0.max_row, 6) + 1):
                row = {norm(ws0.cell(r, c).value) for c in range(1, min(ws0.max_column, 30) + 1)}
                if norm('Gereksinim Numarası') in row:
                    is_srs = True
                    break
            if is_srs:
                reqs_f = reqs_f or f
            else:
                tests_cands.append(f)
            wb.close()
    tests_f = tests_cands[0] if tests_cands else None
    return reqs_f, docx_f, tests_f

# ---------- SRS ----------
def parse_srs(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hr = find_header_row(ws, ['Gereksinim Numarası', 'Gereksinim Numarasi'])
    cm = col_map(ws, hr)
    c_no = pick(cm, 'Gereksinim Numarası', 'Gereksinim No')
    c_ac = pick(cm, 'Gereksinim Açıklaması', 'Açıklama')
    c_kr = pick(cm, 'Kritiklik')
    c_on = pick(cm, 'Öncelik')
    c_mod = pick(cm, 'Modül', 'Modul')
    c_ks = pick(cm, 'Kullanım Senaryo Numarası', 'Kullanim Senaryo Numarasi')
    srs = {}
    for r in range(hr + 1, ws.max_row + 1):
        no = ws.cell(r, c_no).value if c_no else None
        if not no or not str(no).strip().upper().startswith('SRS'):
            continue
        no = str(no).strip()
        srs[no] = {
            'aciklama': (ws.cell(r, c_ac).value if c_ac else '') or '',
            'kritiklik': (ws.cell(r, c_kr).value if c_kr else '') or '',
            'oncelik': (ws.cell(r, c_on).value if c_on else '') or '',
            'modul': (ws.cell(r, c_mod).value if c_mod else '') or '',
            'ks': (ws.cell(r, c_ks).value if c_ks else '') or '',
        }
    return srs

# ---------- Kullanım senaryoları (docx) ----------
LABELS = {
    'no': ['Kullanım Senaryo Numarası', 'Kullanim Senaryo Numarasi'],
    'ad': ['Kullanım Senaryo Adı', 'Kullanim Senaryo Adi'],
    'gereksinim': ['İlgili Gereksinimler', 'Ilgili Gereksinimler'],
    'aktor': ['Aktörler', 'Aktorler'],
    'onkosul': ['Ön Koşul', 'On Kosul'],
    'sonkosul': ['Son Koşul', 'Son Kosul'],
    'ana': ['Başarılı Ana Senaryo', 'Basarili Ana Senaryo'],
    'alt': ['Alternatif Senaryolar', 'Alternatif Senaryo'],
    'iskural': ['İş Kuralları', 'Is Kurallari'],
}
def label_key(text):
    n = norm(text)
    for k, al in LABELS.items():
        for a in al:
            if n == norm(a) or n.startswith(norm(a)):
                return k
    return None

def parse_usecases(path):
    d = docx.Document(path)
    ucs = []
    for t in d.tables:
        # bu tablo bir kullanım senaryosu mu? ilk kolonda 'Kullanım Senaryo Numarası' geçiyor mu
        is_uc = False
        for row in t.rows[:3]:
            if label_key(row.cells[0].text) == 'no':
                is_uc = True
                break
        if not is_uc:
            continue
        uc = {k: '' for k in LABELS}
        for row in t.rows:
            k = label_key(row.cells[0].text)
            if k:
                uc[k] = row.cells[-1].text.strip()
        if uc['no']:
            ucs.append(uc)
    return ucs

# ---------- Testler (xlsx) ----------
def parse_tests(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        hr = find_header_row(ws, ['ID', 'Test Senaryo Adı', 'Test Senaryo Adi'])
        cm = col_map(ws, hr)
        c_id = pick(cm, 'ID')
        c_ad = pick(cm, 'Test Senaryo Adı', 'Test Senaryo Adi')
        c_ac = pick(cm, 'Test Senaryo Açıklama', 'Test Senaryo Aciklama', 'Açıklama')
        c_on = pick(cm, 'Ön Koşul', 'On Kosul')
        c_ger = pick(cm, 'Gereksinim')
        c_adim = pick(cm, 'Test Adımları', 'Test Adimlari')
        c_bek = pick(cm, 'Beklenen Sonuç', 'Beklenen Sonuc')
        c_ks = pick(cm, 'Kullanım Senaryo Numarası', 'Kullanim Senaryo Numarasi')
        if not c_id:
            continue
        order, tests = [], {}
        cur = None
        id_hdr = norm(ws.cell(hr, c_id).value)  # tekrar eden başlık satırlarını ele
        for r in range(hr + 1, ws.max_row + 1):
            rid = ws.cell(r, c_id).value
            if rid and str(rid).strip() and norm(rid) not in (id_hdr, 'id'):
                cur = str(rid).strip()
                if cur not in tests:
                    tests[cur] = {'ad': ws.cell(r, c_ad).value if c_ad else '',
                                  'aciklama': ws.cell(r, c_ac).value if c_ac else '',
                                  'onkosul': ws.cell(r, c_on).value if c_on else '',
                                  'gereksinim': ws.cell(r, c_ger).value if c_ger else '',
                                  'ks': ws.cell(r, c_ks).value if c_ks else '',
                                  'adimlar': [], 'beklenen': []}
                    order.append(cur)
            if cur is None:
                continue
            a = ws.cell(r, c_adim).value if c_adim else None
            b = ws.cell(r, c_bek).value if c_bek else None
            if a: tests[cur]['adimlar'].append(str(a).strip())
            if b: tests[cur]['beklenen'].append(str(b).strip())
        if order:
            sheets[name] = (order, tests)
    return sheets

# ---------- bundle yazımı ----------
def write_bundle(outdir, ks, uc, reqs, srs, sheetname, order, tests):
    L = []
    L.append(f"# {uc['no']} — {uc['ad']}\n")
    L.append("## KULLANIM SENARYOSU\n")
    L.append(f"**Aktörler:** {uc['aktor']}\n")
    L.append("### Ön Koşul\n" + uc['onkosul'] + "\n")
    L.append("### Son Koşul\n" + uc['sonkosul'] + "\n")
    L.append("### Başarılı Ana Senaryo (adımlar)\n")
    for i, s in enumerate([x for x in uc['ana'].split('\n') if x.strip()], 1):
        L.append(f"{i}. {s.strip()}")
    L.append("\n### Alternatif Senaryolar\n")
    for i, s in enumerate([x for x in uc['alt'].split('\n') if x.strip()], 1):
        L.append(f"{i}. {s.strip()}")
    L.append("\n### İş Kuralları\n")
    for i, s in enumerate([x for x in uc['iskural'].split('\n') if x.strip()], 1):
        L.append(f"{i}. {s.strip()}")
    L.append("\n## İLGİLİ SRS GEREKSİNİMLERİ\n")
    for rq in reqs:
        info = srs.get(rq, {})
        L.append(f"- **{rq}** (Kritiklik:{info.get('kritiklik')}, Öncelik:{info.get('oncelik')}): {info.get('aciklama')}")
    L.append(f"\n## TEST SENARYOLARI (sayfa {sheetname}, toplam {len(order)} test)\n")
    for tid in order:
        t = tests[tid]
        L.append(f"### {tid} — {t['ad']}")
        L.append(f"- Açıklama: {t['aciklama']}")
        L.append(f"- Ön Koşul: {t['onkosul']}")
        L.append(f"- Gereksinim etiketi: {t['gereksinim'] if t['gereksinim'] else '(BOŞ)'}")
        L.append(f"- Test Adımları: {' | '.join(t['adimlar'])}")
        L.append(f"- Beklenen Sonuç: {' | '.join(t['beklenen'])}")
        L.append("")
    with open(os.path.join(outdir, f"{ks}_input.md"), 'w', encoding='utf-8') as f:
        f.write('\n'.join(L))

def match_sheet(ks, sheets):
    if ks in sheets:
        return ks
    for name in sheets:
        if ks_short(name) == ks or ks in name:
            return name
    return None

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--indir')
    ap.add_argument('--reqs'); ap.add_argument('--usecases'); ap.add_argument('--tests')
    ap.add_argument('--outdir', default='work')
    ap.add_argument('--module', default=None)
    a = ap.parse_args()

    reqs_f, docx_f, tests_f = a.reqs, a.usecases, a.tests
    if a.indir:
        r, d, t = classify(a.indir)
        reqs_f = reqs_f or r; docx_f = docx_f or d; tests_f = tests_f or t
    if not (reqs_f and docx_f and tests_f):
        print(f"HATA: 3 kaynak da bulunamadı. SRS={reqs_f} docx={docx_f} test={tests_f}", file=sys.stderr)
        sys.exit(2)
    os.makedirs(a.outdir, exist_ok=True)
    print(f"SRS={os.path.basename(reqs_f)}\ndocx={os.path.basename(docx_f)}\ntest={os.path.basename(tests_f)}")

    srs = parse_srs(reqs_f)
    ucs = parse_usecases(docx_f)
    sheets = parse_tests(tests_f)

    module = a.module or (module_of(ucs[0]['no']) if ucs else None) or 'MOD'
    srs_universe = sorted([k for k, v in srs.items()
                           if (not v.get('modul')) or norm(v['modul']) == norm(module)],
                          key=lambda s: int(re.findall(r'\d+', s)[-1]))
    if not srs_universe:
        srs_universe = sorted(srs.keys(), key=lambda s: int(re.findall(r'\d+', s)[-1]))

    manifest = {'module': module, 'tests_file': os.path.abspath(tests_f),
                'srs_universe': srs_universe, 'use_cases': {}}
    for uc in ucs:
        ks = ks_short(uc['no'])
        reqs = [x.strip() for x in uc['gereksinim'].split('\n') if x.strip().upper().startswith('SRS')]
        sn = match_sheet(ks, sheets)
        if not sn:
            print(f"UYARI: {ks} için test sayfası bulunamadı, atlanıyor.", file=sys.stderr)
            continue
        order, tests = sheets[sn]
        write_bundle(a.outdir, ks, uc, reqs, srs, sn, order, tests)
        manifest['use_cases'][ks] = {'ad': uc['ad'], 'sheet': sn, 'srs': reqs,
                                     'valid_test_ids': order}
        print(f"{ks}: {len(reqs)} SRS, {len(order)} test -> {a.outdir}/{ks}_input.md")
    with open(os.path.join(a.outdir, 'manifest.json'), 'w', encoding='utf-8') as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    print(f"\nmodül={module} · {len(manifest['use_cases'])} KS · {len(srs_universe)} SRS · manifest.json yazıldı")

if __name__ == '__main__':
    main()
